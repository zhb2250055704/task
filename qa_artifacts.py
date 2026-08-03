#!/usr/bin/env python3
"""Generate XMind and Excel deliverables from structured QA design data."""

import io
import json
import re
import uuid
import zipfile
from collections import OrderedDict


XMIND_MIME = 'application/vnd.xmind.workbook'
XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _text(value):
    if value is None:
        return ''
    if isinstance(value, (list, tuple, set)):
        return '、'.join(_text(item) for item in value if _text(item))
    return str(value).strip()


def _safe_filename(value, fallback):
    name = re.sub(r'[\\/:*?"<>|\x00-\x1f]+', '-', _text(value)).strip(' .-')
    return (name or fallback)[:80]


def _topic(title, children=None, notes=''):
    topic = {
        'id': uuid.uuid4().hex,
        'class': 'topic',
        'title': _text(title) or '未命名测试点',
    }
    if children:
        topic['children'] = {'attached': children}
    if notes:
        topic['notes'] = {'plain': {'content': _text(notes)}}
    return topic


def _items(value):
    if isinstance(value, (list, tuple, set)):
        return [_text(item) for item in value if _text(item)]
    text = _text(value)
    return [text] if text else []


def _section_topic(title, values, fallback='无'):
    items = _items(values)
    return _topic(title, [_topic(item) for item in items] or [_topic(fallback)])


def _point_steps(point):
    steps = []
    for item in point.get('steps', []):
        if not isinstance(item, dict):
            continue
        action = _text(item.get('action'))
        expected = _text(item.get('expected'))
        if action or expected:
            steps.append({'action': action or '执行测试操作', 'expected': expected or '预期结果待确认'})
    if steps:
        return steps

    action = _text(point.get('scenario')) or '执行测试场景'
    expected = _text(point.get('target')) or _text(point.get('expected_results')) or '预期结果待确认'
    return [{'action': action, 'expected': expected}]


def build_test_points_xmind(design, title=''):
    points = [item for item in design.get('test_points', []) if isinstance(item, dict)]
    if not points:
        raise ValueError('测试点结果为空，无法生成 XMind 文件')

    modules = OrderedDict()
    for point in points:
        module = _text(point.get('module')) or '未分类'
        feature = _text(point.get('feature')) or '核心功能'
        dimension = _text(point.get('dimension')) or _text(point.get('type')) or '功能验证'
        modules.setdefault(module, OrderedDict()).setdefault(feature, OrderedDict()).setdefault(
            dimension, []
        ).append(point)

    module_topics = []
    for module, feature_groups in modules.items():
        feature_topics = []
        module_count = 0
        for feature, dimension_groups in feature_groups.items():
            dimension_topics = []
            for dimension, dimension_points in dimension_groups.items():
                point_topics = []
                module_count += len(dimension_points)
                for point in dimension_points:
                    point_id = _text(point.get('id'))
                    priority = _text(point.get('priority'))
                    point_type = _text(point.get('type'))
                    prefix = ''.join(f'[{value}]' for value in (point_id, priority, point_type) if value)
                    steps = _point_steps(point)
                    step_topics = [
                        _topic(
                            f'{index}. {step["action"]}',
                            [_topic(f'预期：{step["expected"]}')],
                        )
                        for index, step in enumerate(steps, 1)
                    ]
                    expected_results = _items(point.get('expected_results'))
                    if not expected_results:
                        expected_results = list(dict.fromkeys(
                            step['expected'] for step in steps if step.get('expected')
                        ))
                    details = [
                        _section_topic(
                            '前置条件',
                            point.get('preconditions') or point.get('precondition'),
                        ),
                        _section_topic('测试数据', point.get('test_data'), '按场景准备对应数据'),
                        _topic('操作步骤', step_topics),
                        _section_topic('最终检查', expected_results, '预期结果待确认'),
                    ]
                    notes = '\n'.join(filter(None, (
                        f'来源：{_text(point.get("source"))}' if _text(point.get('source')) else '',
                        f'需求追踪：{_text(point.get("requirement_ids"))}'
                        if _text(point.get('requirement_ids')) else '',
                    )))
                    point_topics.append(_topic(
                        f'{prefix} {_text(point.get("scenario"))}'.strip(),
                        details,
                        notes=notes,
                    ))
                dimension_topics.append(_topic(dimension, point_topics))
            feature_topics.append(_topic(feature, dimension_topics))
        module_topics.append(_topic(f'{module}（{module_count}）', feature_topics))

    document_title = _text(title) or _text(design.get('title')) or '测试点设计'
    summary = design.get('summary') if isinstance(design.get('summary'), dict) else {}
    overview = _topic('测试概览', [
        _topic(f'测试点数量：{len(points)}'),
        _topic(f'测试目标：{_text(summary.get("objective")) or "待确认"}'),
        _topic(f'覆盖范围：{_text(summary.get("scope")) or "待确认"}'),
        _topic(f'最高风险：{_text(summary.get("highest_risk")) or "待确认"}'),
    ])
    sheet_id = uuid.uuid4().hex
    content = [{
        'id': sheet_id,
        'class': 'sheet',
        'title': '测试点',
        'rootTopic': {
            'id': uuid.uuid4().hex,
            'class': 'topic',
            'title': document_title,
            'structureClass': 'org.xmind.ui.logic.right',
            'children': {'attached': [overview, *module_topics]},
        },
    }]
    metadata = {
        'creator': {'name': 'GM 命令管理工具', 'version': '1.0'},
        'activeSheetId': sheet_id,
    }
    manifest = {
        'file-entries': {
            'content.json': {'media-type': 'application/json'},
            'metadata.json': {'media-type': 'application/json'},
        },
    }

    output = io.BytesIO()
    with zipfile.ZipFile(output, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr('content.json', json.dumps(content, ensure_ascii=False, separators=(',', ':')))
        archive.writestr('metadata.json', json.dumps(metadata, ensure_ascii=False, separators=(',', ':')))
        archive.writestr('manifest.json', json.dumps(manifest, ensure_ascii=False, separators=(',', ':')))
    return {
        'content': output.getvalue(),
        'filename': _safe_filename(document_title, '测试点') + '-测试点.xmind',
        'mime': XMIND_MIME,
        'format': 'xmind',
        'count': len(points),
    }


def _clean_cell(value):
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', _text(value))


def _style_sheet(sheet, widths, priority_column=None):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill('solid', fgColor='315C4D')
    header_font = Font(color='FFFFFF', bold=True)
    stripe_fill = PatternFill('solid', fgColor='F1F6F3')
    p0_fill = PatternFill('solid', fgColor='FDE9E7')
    p1_fill = PatternFill('solid', fgColor='FFF2D8')
    border = Border(bottom=Side(style='thin', color='D7E2DC'))

    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 28
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row_index, row in enumerate(sheet.iter_rows(min_row=2), 2):
        sheet.row_dimensions[row_index].height = 44
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border
            if row_index % 2 == 0:
                cell.fill = stripe_fill
        if priority_column:
            priority = _text(sheet.cell(row=row_index, column=priority_column).value).upper()
            if priority == 'P0':
                sheet.cell(row=row_index, column=priority_column).fill = p0_fill
            elif priority == 'P1':
                sheet.cell(row=row_index, column=priority_column).fill = p1_fill
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width


def build_test_cases_xlsx(design, title=''):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError('缺少 Excel 生成组件 openpyxl') from exc

    cases = [item for item in design.get('test_cases', []) if isinstance(item, dict)]
    if not cases:
        raise ValueError('测试用例结果为空，无法生成 Excel 文件')

    workbook = Workbook()
    cases_sheet = workbook.active
    cases_sheet.title = '测试用例'
    case_headers = [
        '用例ID', '模块', '用例标题', '优先级', '测试类型', '关联需求', '关联测试点',
        '前置条件', '测试数据', '测试步骤', '预期结果', '自动化适合度',
    ]
    cases_sheet.append(case_headers)
    for case in cases:
        steps = [item for item in case.get('steps', []) if isinstance(item, dict)]
        actions = '\n'.join(
            f'{index}. {_clean_cell(step.get("action"))}' for index, step in enumerate(steps, 1)
        )
        expected = '\n'.join(
            f'{index}. {_clean_cell(step.get("expected"))}' for index, step in enumerate(steps, 1)
        )
        cases_sheet.append([
            _clean_cell(case.get('id')),
            _clean_cell(case.get('module')),
            _clean_cell(case.get('title')),
            _clean_cell(case.get('priority')),
            _clean_cell(case.get('type')),
            _clean_cell(case.get('requirement_ids')),
            _clean_cell(case.get('test_point_ids')),
            _clean_cell(case.get('preconditions')),
            _clean_cell(case.get('test_data')),
            actions,
            expected,
            _clean_cell(case.get('automation')),
        ])
    cases_sheet.sheet_properties.tabColor = '315C4D'
    _style_sheet(cases_sheet, [14, 18, 42, 10, 12, 18, 18, 32, 30, 52, 52, 14], priority_column=4)

    steps_sheet = workbook.create_sheet('测试步骤')
    steps_sheet.append(['用例ID', '步骤序号', '操作步骤', '预期结果'])
    for case in cases:
        for index, step in enumerate(case.get('steps', []), 1):
            if not isinstance(step, dict):
                continue
            steps_sheet.append([
                _clean_cell(case.get('id')),
                index,
                _clean_cell(step.get('action')),
                _clean_cell(step.get('expected')),
            ])
    steps_sheet.sheet_properties.tabColor = '5F8C78'
    _style_sheet(steps_sheet, [14, 10, 58, 58])

    trace_sheet = workbook.create_sheet('需求追踪')
    trace_sheet.append(['需求ID', '关联测试点', '关联测试用例', '覆盖状态'])
    for item in design.get('traceability', []):
        if not isinstance(item, dict):
            continue
        trace_sheet.append([
            _clean_cell(item.get('requirement_id')),
            _clean_cell(item.get('test_point_ids')),
            _clean_cell(item.get('test_case_ids')),
            _clean_cell(item.get('coverage')),
        ])
    trace_sheet.sheet_properties.tabColor = '8AA899'
    _style_sheet(trace_sheet, [16, 36, 36, 14])

    output = io.BytesIO()
    workbook.save(output)
    document_title = _text(title) or _text(design.get('title')) or '测试用例'
    return {
        'content': output.getvalue(),
        'filename': _safe_filename(document_title, '测试用例') + '-测试用例.xlsx',
        'mime': XLSX_MIME,
        'format': 'xlsx',
        'count': len(cases),
    }


def build_qa_artifact(design, mode, title=''):
    if not isinstance(design, dict):
        raise ValueError('测试设计结果不是结构化数据')
    if mode == 'points':
        return build_test_points_xmind(design, title)
    if mode == 'cases':
        return build_test_cases_xlsx(design, title)
    raise ValueError('仅支持生成测试点 XMind 或测试用例 Excel')
