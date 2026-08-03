import io
import json
import unittest
import zipfile

from openpyxl import load_workbook

from qa_artifacts import build_qa_artifact
from qa_local_engine import build_rule_design, parse_qa_design_json


SAMPLE_DESIGN = {
    'title': '账号登录需求',
    'summary': {
        'objective': '验证账号登录功能',
        'scope': ['登录', '账号'],
        'highest_risk': '账号串号',
    },
    'requirements': [
        {'id': 'REQ-X', 'module': '登录', 'behavior': '有效账号可以登录'},
    ],
    'test_points': [
        {
            'id': 'POINT-X',
            'requirement_ids': ['REQ-X'],
            'module': '登录',
            'feature': '账号登录',
            'dimension': '主流程',
            'preconditions': ['账号已创建', '登录服务可用'],
            'test_data': ['有效账号 A', '正确密码 P'],
            'scenario': '使用有效账号登录',
            'type': '正向',
            'priority': 'P0',
            'steps': [
                {'action': '输入账号 A 和密码 P，点击登录', 'expected': '登录请求成功，进入角色界面'},
                {'action': '查看当前角色信息', 'expected': '显示账号 A 对应的角色信息'},
            ],
            'expected_results': ['登录成功并进入角色界面', '当前角色属于账号 A'],
            'source': '需求.docx',
        },
    ],
    'test_cases': [
        {
            'id': 'CASE-X',
            'requirement_ids': ['REQ-X'],
            'test_point_ids': ['POINT-X'],
            'module': '登录',
            'title': '有效账号登录',
            'preconditions': ['账号已创建'],
            'test_data': ['有效账号 A'],
            'steps': [
                {'action': '输入账号并登录', 'expected': '进入角色界面'},
            ],
            'priority': 'P0',
            'type': '正向',
            'automation': '高',
        },
    ],
    'traceability': [
        {
            'requirement_id': 'REQ-X',
            'test_point_ids': ['POINT-X'],
            'test_case_ids': ['CASE-X'],
            'coverage': '已覆盖',
        },
    ],
}


class QaArtifactTests(unittest.TestCase):
    def test_xmind_contains_structured_test_points(self):
        artifact = build_qa_artifact(SAMPLE_DESIGN, 'points', '账号登录需求')
        self.assertEqual(artifact['format'], 'xmind')
        self.assertEqual(artifact['count'], 1)
        with zipfile.ZipFile(io.BytesIO(artifact['content'])) as archive:
            self.assertTrue({'content.json', 'metadata.json', 'manifest.json'} <= set(archive.namelist()))
            content = json.loads(archive.read('content.json').decode('utf-8'))
        root = content[0]['rootTopic']
        self.assertEqual(root['title'], '账号登录需求')
        serialized = json.dumps(content, ensure_ascii=False)
        self.assertIn('使用有效账号登录', serialized)
        self.assertIn('登录成功并进入角色界面', serialized)
        self.assertIn('登录（1）', serialized)
        self.assertIn('账号登录', serialized)
        self.assertIn('主流程', serialized)
        self.assertIn('前置条件', serialized)
        self.assertIn('账号已创建', serialized)
        self.assertIn('测试数据', serialized)
        self.assertIn('有效账号 A', serialized)
        self.assertIn('操作步骤', serialized)
        self.assertIn('输入账号 A 和密码 P，点击登录', serialized)
        self.assertIn('预期：登录请求成功，进入角色界面', serialized)
        self.assertIn('最终检查', serialized)
        self.assertNotIn('关联需求：', serialized)

    def test_xlsx_contains_cases_steps_and_traceability(self):
        artifact = build_qa_artifact(SAMPLE_DESIGN, 'cases', '账号登录需求')
        self.assertEqual(artifact['format'], 'xlsx')
        workbook = load_workbook(io.BytesIO(artifact['content']))
        self.assertEqual(workbook.sheetnames, ['测试用例', '测试步骤', '需求追踪'])
        self.assertEqual(workbook['测试用例']['A2'].value, 'CASE-X')
        self.assertIn('输入账号并登录', workbook['测试用例']['J2'].value)
        self.assertEqual(workbook['测试步骤']['D2'].value, '进入角色界面')
        self.assertEqual(workbook['需求追踪']['D2'].value, '已覆盖')

    def test_codex_json_is_normalized_for_artifact_generation(self):
        raw = '```json\n' + json.dumps(SAMPLE_DESIGN, ensure_ascii=False) + '\n```'
        parsed = parse_qa_design_json(raw, mode='cases', title='账号登录需求')
        self.assertEqual(parsed['requirements'][0]['id'], 'REQ-001')
        self.assertEqual(parsed['test_points'][0]['id'], 'TP-001')
        self.assertEqual(parsed['test_points'][0]['feature'], '账号登录')
        self.assertEqual(parsed['test_points'][0]['dimension'], '主流程')
        self.assertEqual(parsed['test_points'][0]['preconditions'][0], '账号已创建')
        self.assertEqual(parsed['test_points'][0]['steps'][0]['expected'], '登录请求成功，进入角色界面')
        self.assertEqual(parsed['test_cases'][0]['id'], 'TC-001')
        self.assertEqual(parsed['traceability'][0]['coverage'], '已覆盖')

    def test_legacy_test_point_is_upgraded_to_executable_structure(self):
        legacy = json.loads(json.dumps(SAMPLE_DESIGN, ensure_ascii=False))
        legacy['test_points'][0] = {
            'requirement_ids': ['REQ-X'],
            'module': '登录',
            'precondition': '账号已创建',
            'scenario': '使用有效账号登录',
            'type': '正向',
            'priority': 'P0',
            'target': '登录成功并进入角色界面',
        }
        parsed = parse_qa_design_json(json.dumps(legacy, ensure_ascii=False), mode='points')
        point = parsed['test_points'][0]
        self.assertEqual(point['preconditions'], ['账号已创建'])
        self.assertEqual(point['expected_results'], ['登录成功并进入角色界面'])
        self.assertEqual(point['steps'][0], {
            'action': '使用有效账号登录',
            'expected': '登录成功并进入角色界面',
        })

    def test_rule_engine_points_are_directly_executable(self):
        design = build_rule_design(
            '活动商城礼包持续 7 天，每日刷新，购买成功后通过邮件发放奖励。',
            [],
            'points',
            'gm',
            'standard',
            '活动商城礼包',
        )
        self.assertTrue(design['test_points'])
        for point in design['test_points']:
            self.assertTrue(point.get('feature'))
            self.assertTrue(point.get('dimension'))
            self.assertTrue(point.get('preconditions'))
            self.assertTrue(point.get('test_data'))
            self.assertTrue(point.get('steps'))
            self.assertTrue(point.get('expected_results'))
            self.assertTrue(all(step.get('action') and step.get('expected') for step in point['steps']))


if __name__ == '__main__':
    unittest.main()
